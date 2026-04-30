"""
SQLite database layer for DiceBot.
Single file: ~/.dicebot-ticketkings/dicebot.db
Thread-safe with WAL mode for concurrent reads.
"""

from __future__ import annotations

import json
import os
import sqlite3
import time
import threading
from datetime import datetime, UTC

DB_DIR = os.path.join(os.path.expanduser("~"), ".dicebot-ticketkings")
DB_PATH = os.path.join(DB_DIR, "dicebot.db")
SESSION_MAX_AGE = 6 * 24 * 3600  # 6 days
SESSION_WARN_AGE = 5 * 24 * 3600  # 5 days — "expiring soon"

_lock = threading.Lock()


def _connect() -> sqlite3.Connection:
    os.makedirs(DB_DIR, exist_ok=True)
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn


def _ensure_columns(conn: sqlite3.Connection, table: str, columns: dict[str, str]) -> None:
    existing = {
        row["name"]
        for row in conn.execute(f"PRAGMA table_info({table})").fetchall()
    }
    for name, ddl in columns.items():
        if name in existing:
            continue
        conn.execute(f"ALTER TABLE {table} ADD COLUMN {name} {ddl}")


def init_db() -> None:
    """Create tables if they don't exist."""
    with _lock:
        conn = _connect()
        conn.executescript("""
            CREATE TABLE IF NOT EXISTS groups (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                created_at TEXT DEFAULT (datetime('now'))
            );

            CREATE TABLE IF NOT EXISTS accounts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                phone TEXT NOT NULL UNIQUE,
                email TEXT DEFAULT '',
                card_number TEXT DEFAULT '',
                card_exp_month TEXT DEFAULT '',
                card_exp_year TEXT DEFAULT '',
                card_cvv TEXT DEFAULT '',
                billing_name TEXT DEFAULT '',
                billing_email TEXT DEFAULT '',
                billing_phone TEXT DEFAULT '',
                billing_postal TEXT DEFAULT '',
                billing_country TEXT DEFAULT 'US',
                proxy TEXT DEFAULT '',
                aycd_key TEXT DEFAULT '',
                imap_email TEXT DEFAULT '',
                imap_password TEXT DEFAULT '',
                group_id INTEGER REFERENCES groups(id) ON DELETE SET NULL,
                created_at TEXT DEFAULT (datetime('now'))
            );

            CREATE TABLE IF NOT EXISTS sessions (
                account_id INTEGER PRIMARY KEY REFERENCES accounts(id) ON DELETE CASCADE,
                bearer_token TEXT NOT NULL,
                device_id TEXT DEFAULT '',
                saved_at REAL NOT NULL,
                phone TEXT DEFAULT ''
            );

            CREATE TABLE IF NOT EXISTS inventory_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                record_key TEXT NOT NULL UNIQUE,
                purchase_id TEXT DEFAULT '',
                account_id INTEGER REFERENCES accounts(id) ON DELETE SET NULL,
                account_name TEXT DEFAULT '',
                account_phone TEXT DEFAULT '',
                event_url TEXT DEFAULT '',
                event_name TEXT DEFAULT '',
                event_date TEXT DEFAULT '',
                event_venue TEXT DEFAULT '',
                ticket_type_id TEXT DEFAULT '',
                ticket_name TEXT DEFAULT '',
                ticket_currency TEXT DEFAULT 'USD',
                ticket_price REAL DEFAULT 0,
                quantity INTEGER DEFAULT 1,
                total_price REAL DEFAULT 0,
                purchase_status TEXT DEFAULT 'purchased',
                purchased_at TEXT DEFAULT (datetime('now'))
            );

            INSERT OR IGNORE INTO groups (name) VALUES ('Default');

            DROP TABLE IF EXISTS account_cards;

            CREATE TABLE IF NOT EXISTS payment_pools (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                created_at TEXT DEFAULT (datetime('now'))
            );

            CREATE TABLE IF NOT EXISTS payment_cards (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                pool_id INTEGER REFERENCES payment_pools(id) ON DELETE SET NULL,
                card_number TEXT NOT NULL,
                card_exp_month TEXT DEFAULT '',
                card_exp_year TEXT DEFAULT '',
                card_cvv TEXT DEFAULT '',
                billing_name TEXT DEFAULT '',
                billing_email TEXT DEFAULT '',
                billing_phone TEXT DEFAULT '',
                billing_postal TEXT DEFAULT '',
                billing_country TEXT DEFAULT 'US',
                created_at TEXT DEFAULT (datetime('now'))
            );
            CREATE INDEX IF NOT EXISTS idx_payment_cards_pool ON payment_cards(pool_id);

            CREATE TABLE IF NOT EXISTS account_card_assignments (
                account_id INTEGER NOT NULL REFERENCES accounts(id) ON DELETE CASCADE,
                card_id INTEGER NOT NULL REFERENCES payment_cards(id) ON DELETE CASCADE,
                created_at TEXT DEFAULT (datetime('now')),
                PRIMARY KEY (account_id, card_id)
            );
            CREATE INDEX IF NOT EXISTS idx_aca_card ON account_card_assignments(card_id);
            CREATE INDEX IF NOT EXISTS idx_aca_account ON account_card_assignments(account_id);

            CREATE TABLE IF NOT EXISTS code_pools (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                created_at TEXT DEFAULT (datetime('now'))
            );

            CREATE TABLE IF NOT EXISTS code_pool_codes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                pool_id INTEGER NOT NULL REFERENCES code_pools(id) ON DELETE CASCADE,
                code TEXT NOT NULL,
                created_at TEXT DEFAULT (datetime('now'))
            );
            CREATE INDEX IF NOT EXISTS idx_code_pool_codes_pool ON code_pool_codes(pool_id);
            CREATE UNIQUE INDEX IF NOT EXISTS idx_code_pool_codes_unique ON code_pool_codes(pool_id, code);

            CREATE TABLE IF NOT EXISTS proxy_pools (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                created_at TEXT DEFAULT (datetime('now'))
            );

            CREATE TABLE IF NOT EXISTS proxy_pool_proxies (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                pool_id INTEGER NOT NULL REFERENCES proxy_pools(id) ON DELETE CASCADE,
                proxy TEXT NOT NULL,
                created_at TEXT DEFAULT (datetime('now'))
            );
            CREATE INDEX IF NOT EXISTS idx_proxy_pool_proxies_pool ON proxy_pool_proxies(pool_id);
            CREATE UNIQUE INDEX IF NOT EXISTS idx_proxy_pool_proxies_unique ON proxy_pool_proxies(pool_id, proxy);
        """)
        _ensure_columns(conn, "accounts", {
            "imap_email": "TEXT DEFAULT ''",
            "imap_password": "TEXT DEFAULT ''",
        })
        conn.commit()
        conn.close()


# ── Groups ─────────────────────────────────────────────────────────────────

def get_groups() -> list[dict]:
    conn = _connect()
    rows = conn.execute("SELECT * FROM groups ORDER BY name").fetchall()
    conn.close()
    return [dict(r) for r in rows]


def create_group(name: str) -> int:
    with _lock:
        conn = _connect()
        cur = conn.execute("INSERT INTO groups (name) VALUES (?)", (name.strip(),))
        conn.commit()
        gid = cur.lastrowid
        conn.close()
        return gid


def delete_group(group_id: int) -> None:
    with _lock:
        conn = _connect()
        conn.execute("UPDATE accounts SET group_id = NULL WHERE group_id = ?", (group_id,))
        conn.execute("DELETE FROM groups WHERE id = ?", (group_id,))
        conn.commit()
        conn.close()


def rename_group(group_id: int, name: str) -> None:
    with _lock:
        conn = _connect()
        conn.execute("UPDATE groups SET name = ? WHERE id = ?", (name.strip(), group_id))
        conn.commit()
        conn.close()


# ── Payment cards (cards belong to a managed pool, assigned to accounts) ──

_CARD_FIELDS = (
    "card_number", "card_exp_month", "card_exp_year", "card_cvv",
    "billing_name", "billing_email", "billing_phone",
    "billing_postal", "billing_country",
)


def get_payment_pools() -> list[dict]:
    conn = _connect()
    rows = conn.execute("""
        SELECT p.id, p.name, p.created_at,
               (SELECT COUNT(*) FROM payment_cards c WHERE c.pool_id = p.id) AS card_count
        FROM payment_pools p
        ORDER BY p.name
    """).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def create_payment_pool(name: str) -> int:
    name = (name or "").strip()
    if not name:
        raise ValueError("Pool name is required")
    with _lock:
        conn = _connect()
        cur = conn.execute("INSERT INTO payment_pools (name) VALUES (?)", (name,))
        conn.commit()
        pid = cur.lastrowid
        conn.close()
        return int(pid)


def rename_payment_pool(pool_id: int, name: str) -> None:
    name = (name or "").strip()
    if not name:
        raise ValueError("Pool name is required")
    with _lock:
        conn = _connect()
        conn.execute(
            "UPDATE payment_pools SET name = ? WHERE id = ?",
            (name, int(pool_id)),
        )
        conn.commit()
        conn.close()


def delete_payment_pool(pool_id: int) -> None:
    with _lock:
        conn = _connect()
        conn.execute("DELETE FROM payment_pools WHERE id = ?", (int(pool_id),))
        conn.commit()
        conn.close()


def _ensure_pool_id(name: str) -> int:
    """Look up a pool by name (case-insensitive). Auto-create if missing."""
    name = (name or "").strip()
    if not name:
        raise ValueError("Pool name is required")
    conn = _connect()
    row = conn.execute(
        "SELECT id FROM payment_pools WHERE LOWER(name) = LOWER(?)", (name,)
    ).fetchone()
    conn.close()
    if row:
        return int(row["id"])
    return create_payment_pool(name)


def get_payment_cards() -> list[dict]:
    """All cards joined with their pool, plus assigned account list."""
    conn = _connect()
    rows = conn.execute("""
        SELECT c.*, p.name AS label
        FROM payment_cards c
        LEFT JOIN payment_pools p ON p.id = c.pool_id
        ORDER BY p.name, c.id
    """).fetchall()
    cards = [dict(r) for r in rows]
    if not cards:
        conn.close()
        return cards
    assigns = conn.execute(
        "SELECT a.card_id, a.account_id, ac.phone AS account_phone, ac.name AS account_name "
        "FROM account_card_assignments a "
        "INNER JOIN accounts ac ON ac.id = a.account_id"
    ).fetchall()
    conn.close()
    by_card: dict[int, list[dict]] = {}
    for r in assigns:
        by_card.setdefault(int(r["card_id"]), []).append({
            "account_id": int(r["account_id"]),
            "account_phone": r["account_phone"] or "",
            "account_name": r["account_name"] or "",
        })
    for c in cards:
        c["assigned_accounts"] = by_card.get(int(c["id"]), [])
        if c.get("label") is None:
            c["label"] = ""
    return cards


def get_payment_card(card_id: int) -> dict | None:
    conn = _connect()
    row = conn.execute(
        "SELECT c.*, p.name AS label "
        "FROM payment_cards c LEFT JOIN payment_pools p ON p.id = c.pool_id "
        "WHERE c.id = ?",
        (int(card_id),),
    ).fetchone()
    conn.close()
    if not row:
        return None
    d = dict(row)
    if d.get("label") is None:
        d["label"] = ""
    return d


def add_payment_card(pool_id: int | None = None, label: str | None = None, **fields) -> int:
    """Insert a new card. Provide pool_id directly, or label (will be looked up / auto-created)."""
    if pool_id is None and label:
        pool_id = _ensure_pool_id(label)
    if pool_id is None:
        raise ValueError("pool_id or label is required")
    data = {
        k: ("" if fields.get(k) is None else str(fields.get(k)))
        for k in _CARD_FIELDS
    }
    if not data["billing_country"]:
        data["billing_country"] = "US"
    if not data["card_number"]:
        raise ValueError("card_number is required")
    with _lock:
        conn = _connect()
        cur = conn.execute(f"""
            INSERT INTO payment_cards (pool_id, {", ".join(_CARD_FIELDS)})
            VALUES (?, {", ".join("?" for _ in _CARD_FIELDS)})
        """, (int(pool_id), *(data[k] for k in _CARD_FIELDS)))
        conn.commit()
        cid = cur.lastrowid
        conn.close()
        return int(cid)


def update_payment_card(card_id: int, pool_id: int | None = None, label: str | None = None, **fields) -> None:
    sets: list[str] = []
    vals: list = []
    if pool_id is not None:
        sets.append("pool_id = ?")
        vals.append(int(pool_id))
    elif label is not None:
        sets.append("pool_id = ?")
        vals.append(_ensure_pool_id(label) if label else None)
    for k in fields:
        if k in _CARD_FIELDS:
            sets.append(f"{k} = ?")
            vals.append("" if fields[k] is None else str(fields[k]))
    if not sets:
        return
    vals.append(int(card_id))
    with _lock:
        conn = _connect()
        conn.execute(f"UPDATE payment_cards SET {', '.join(sets)} WHERE id = ?", vals)
        conn.commit()
        conn.close()


def delete_payment_card(card_id: int) -> None:
    with _lock:
        conn = _connect()
        conn.execute("DELETE FROM payment_cards WHERE id = ?", (int(card_id),))
        conn.commit()
        conn.close()


def assign_card(account_id: int, card_id: int) -> None:
    with _lock:
        conn = _connect()
        conn.execute(
            "INSERT OR IGNORE INTO account_card_assignments (account_id, card_id) VALUES (?, ?)",
            (int(account_id), int(card_id)),
        )
        conn.commit()
        conn.close()


def unassign_card(account_id: int, card_id: int) -> None:
    with _lock:
        conn = _connect()
        conn.execute(
            "DELETE FROM account_card_assignments WHERE account_id = ? AND card_id = ?",
            (int(account_id), int(card_id)),
        )
        conn.commit()
        conn.close()


def get_card_labels() -> list[str]:
    """All managed pool names, sorted. Cart modal shows these even if a pool has 0 cards."""
    conn = _connect()
    rows = conn.execute("SELECT name FROM payment_pools ORDER BY name").fetchall()
    conn.close()
    return [r["name"] for r in rows]


def get_assigned_cards_for_account(account_id: int) -> list[dict]:
    conn = _connect()
    rows = conn.execute("""
        SELECT c.*, p.name AS label
        FROM payment_cards c
        INNER JOIN account_card_assignments a ON a.card_id = c.id
        LEFT JOIN payment_pools p ON p.id = c.pool_id
        WHERE a.account_id = ?
        ORDER BY p.name, c.id
    """, (int(account_id),)).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def bulk_account_cards_by_label(account_ids: list[int], label: str) -> dict[int, dict]:
    """For each account, find the first assigned card whose pool name matches `label`."""
    if not account_ids:
        return {}
    ids = [int(x) for x in account_ids]
    placeholders = ",".join("?" for _ in ids)
    conn = _connect()
    rows = conn.execute(f"""
        SELECT a.account_id, c.*, p.name AS label
        FROM account_card_assignments a
        INNER JOIN payment_cards c ON c.id = a.card_id
        INNER JOIN payment_pools p ON p.id = c.pool_id
        WHERE a.account_id IN ({placeholders}) AND LOWER(p.name) = LOWER(?)
        ORDER BY a.account_id, c.id
    """, [*ids, str(label or "")]).fetchall()
    conn.close()
    result: dict[int, dict] = {}
    for r in rows:
        d = dict(r)
        aid = int(d.pop("account_id"))
        if aid not in result:
            result[aid] = d
    return result


def _normalize_phone_for_match(s: str) -> str:
    """Reduce a phone number to its last 10 digits for matching.

    Handles `+13306346924`, `3306346924`, `(330) 634-6924`, `330-634-6924` etc.
    Last 10 digits is the right grain for US/CA — country code differences
    don't break the match. Returns "" if there aren't 10 digits available, in
    which case we skip matching (better no-match than a false positive).
    """
    if not s:
        return ""
    digits = "".join(ch for ch in str(s) if ch.isdigit())
    return digits[-10:] if len(digits) >= 10 else ""


def _find_account_id_for_card(billing_phone: str, billing_email: str) -> tuple[int | None, int]:
    """Return (account_id, match_count). account_id is set only when exactly
    one account matches; match_count tells the caller why nothing was assigned
    (0 = no match, 2+ = ambiguous so we refuse to guess).
    """
    norm_phone = _normalize_phone_for_match(billing_phone)
    norm_email = (billing_email or "").strip().lower()
    if not norm_phone and not norm_email:
        return None, 0

    conn = _connect()
    matches: list[int] = []
    try:
        # Phone match wins (more specific). The LIKE pattern here is a cheap
        # filter on the normalized last-10 digits — cheaper than re-normalizing
        # every account.phone in Python.
        if norm_phone:
            rows = conn.execute(
                "SELECT id, phone FROM accounts WHERE phone LIKE ?",
                (f"%{norm_phone}",),
            ).fetchall()
            for r in rows:
                if _normalize_phone_for_match(r["phone"]) == norm_phone:
                    matches.append(int(r["id"]))
        if not matches and norm_email:
            rows = conn.execute(
                "SELECT id FROM accounts WHERE LOWER(email) = ?",
                (norm_email,),
            ).fetchall()
            matches = [int(r["id"]) for r in rows]
    finally:
        conn.close()

    unique = list(dict.fromkeys(matches))
    if len(unique) == 1:
        return unique[0], 1
    return None, len(unique)


def bulk_add_payment_cards(rows: list[dict], auto_assign: bool = False) -> dict:
    """Import card rows. Each row creates a fresh card. The 'label' column names a pool
    (auto-created if it doesn't exist).

    If auto_assign is True, after each card is created the importer looks up an
    account whose phone (last 10 digits) matches billing_phone OR whose email
    (case-insensitive) matches billing_email, and assigns the card to it. A
    card is only assigned when exactly one account matches; multi-match rows
    are reported as `ambiguous` so the user can resolve them by hand.

    Returns {added, errors, assigned, unmatched, ambiguous} where the latter
    three are populated only when auto_assign is on.
    """
    added = 0
    assigned = 0
    unmatched: list[dict] = []
    ambiguous: list[dict] = []
    errors: list[dict] = []
    for idx, row in enumerate(rows or []):
        cleaned = {
            str(k).strip().lower().replace(" ", "").replace("_", ""): str(v).strip()
            for k, v in (row or {}).items()
            if v is not None
        }
        card_number = _col(cleaned, "cardnumber", "card")
        if not card_number:
            errors.append({"row": idx + 1, "error": "Missing card_number"})
            continue
        label = _col(cleaned, "label", "pool", "poolname")
        if not label:
            errors.append({"row": idx + 1, "error": "Missing label / pool"})
            continue
        billing_email = _col(cleaned, "billingemail", "email")
        billing_phone = _col(cleaned, "billingphone", "phone", "phonenumber")
        try:
            card_id = add_payment_card(
                label=label,
                card_number=card_number,
                card_exp_month=_col(cleaned, "cardexpmonth", "expmonth"),
                card_exp_year=_col(cleaned, "cardexpyear", "expyear"),
                card_cvv=_col(cleaned, "cvv", "cvc"),
                billing_name=_col(cleaned, "billingname", "fullname", "nameoncard"),
                billing_email=billing_email,
                billing_phone=billing_phone,
                billing_postal=_col(
                    cleaned,
                    "billingpostal", "billingpostalcode", "billingzip",
                    "postal", "zip", "zipcode", "postalcode",
                ),
                billing_country=_country(_col(cleaned, "country", "billingcountry") or "US"),
            )
            added += 1
        except Exception as exc:
            errors.append({"row": idx + 1, "error": str(exc)})
            continue

        if auto_assign:
            account_id, match_count = _find_account_id_for_card(billing_phone, billing_email)
            if account_id is not None:
                assign_card(account_id, card_id)
                assigned += 1
            elif match_count == 0:
                unmatched.append({"row": idx + 1, "billing_phone": billing_phone, "billing_email": billing_email})
            else:
                ambiguous.append({"row": idx + 1, "match_count": match_count, "billing_phone": billing_phone, "billing_email": billing_email})

    return {
        "added": added,
        "errors": errors,
        "assigned": assigned,
        "unmatched": unmatched,
        "ambiguous": ambiguous,
    }


# ── Code pools (presale / access codes) ───────────────────────────────────

def get_code_pools() -> list[dict]:
    conn = _connect()
    rows = conn.execute("""
        SELECT p.id, p.name, p.created_at,
               (SELECT COUNT(*) FROM code_pool_codes c WHERE c.pool_id = p.id) AS code_count
        FROM code_pools p
        ORDER BY p.name
    """).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def create_code_pool(name: str) -> int:
    name = (name or "").strip()
    if not name:
        raise ValueError("Pool name is required")
    with _lock:
        conn = _connect()
        cur = conn.execute("INSERT INTO code_pools (name) VALUES (?)", (name,))
        conn.commit()
        pid = cur.lastrowid
        conn.close()
        return int(pid)


def rename_code_pool(pool_id: int, name: str) -> None:
    name = (name or "").strip()
    if not name:
        raise ValueError("Pool name is required")
    with _lock:
        conn = _connect()
        conn.execute(
            "UPDATE code_pools SET name = ? WHERE id = ?",
            (name, int(pool_id)),
        )
        conn.commit()
        conn.close()


def delete_code_pool(pool_id: int) -> None:
    with _lock:
        conn = _connect()
        conn.execute("DELETE FROM code_pools WHERE id = ?", (int(pool_id),))
        conn.commit()
        conn.close()


def get_code_pool_codes(pool_id: int) -> list[dict]:
    conn = _connect()
    rows = conn.execute(
        "SELECT id, pool_id, code, created_at FROM code_pool_codes WHERE pool_id = ? ORDER BY id",
        (int(pool_id),),
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def add_code_pool_codes(pool_id: int, codes: list[str]) -> dict:
    """Insert a batch of codes into the pool. Skips blanks and duplicates inside the same pool."""
    cleaned: list[str] = []
    seen: set[str] = set()
    for raw in codes or []:
        c = str(raw or "").strip()
        if not c or c in seen:
            continue
        seen.add(c)
        cleaned.append(c)
    if not cleaned:
        return {"added": 0, "skipped": 0}
    added = 0
    skipped = 0
    with _lock:
        conn = _connect()
        for c in cleaned:
            try:
                conn.execute(
                    "INSERT INTO code_pool_codes (pool_id, code) VALUES (?, ?)",
                    (int(pool_id), c),
                )
                added += 1
            except sqlite3.IntegrityError:
                skipped += 1
        conn.commit()
        conn.close()
    return {"added": added, "skipped": skipped}


def delete_code_pool_code(code_id: int) -> None:
    with _lock:
        conn = _connect()
        conn.execute("DELETE FROM code_pool_codes WHERE id = ?", (int(code_id),))
        conn.commit()
        conn.close()


def clear_code_pool(pool_id: int) -> None:
    with _lock:
        conn = _connect()
        conn.execute("DELETE FROM code_pool_codes WHERE pool_id = ?", (int(pool_id),))
        conn.commit()
        conn.close()


def draw_codes_from_pool(pool_id: int, count: int) -> list[str]:
    """Pick up to `count` random codes from the pool. Codes are not consumed (caller can choose to)."""
    n = max(0, int(count or 0))
    if n <= 0:
        return []
    conn = _connect()
    rows = conn.execute(
        "SELECT code FROM code_pool_codes WHERE pool_id = ? ORDER BY RANDOM() LIMIT ?",
        (int(pool_id), n),
    ).fetchall()
    conn.close()
    return [r["code"] for r in rows]


# ── Proxy pools (auth farm proxy groups) ──────────────────────────────────

def get_proxy_pools() -> list[dict]:
    conn = _connect()
    rows = conn.execute("""
        SELECT p.id, p.name, p.created_at,
               (SELECT COUNT(*) FROM proxy_pool_proxies pp WHERE pp.pool_id = p.id) AS proxy_count
        FROM proxy_pools p
        ORDER BY p.name
    """).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def create_proxy_pool(name: str) -> int:
    name = (name or "").strip()
    if not name:
        raise ValueError("Pool name is required")
    with _lock:
        conn = _connect()
        cur = conn.execute("INSERT INTO proxy_pools (name) VALUES (?)", (name,))
        conn.commit()
        pid = cur.lastrowid
        conn.close()
        return int(pid)


def rename_proxy_pool(pool_id: int, name: str) -> None:
    name = (name or "").strip()
    if not name:
        raise ValueError("Pool name is required")
    with _lock:
        conn = _connect()
        conn.execute(
            "UPDATE proxy_pools SET name = ? WHERE id = ?",
            (name, int(pool_id)),
        )
        conn.commit()
        conn.close()


def delete_proxy_pool(pool_id: int) -> None:
    with _lock:
        conn = _connect()
        conn.execute("DELETE FROM proxy_pools WHERE id = ?", (int(pool_id),))
        conn.commit()
        conn.close()


def get_proxy_pool_proxies(pool_id: int) -> list[dict]:
    conn = _connect()
    rows = conn.execute(
        "SELECT id, pool_id, proxy, created_at FROM proxy_pool_proxies WHERE pool_id = ? ORDER BY id",
        (int(pool_id),),
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def add_proxy_pool_proxies(pool_id: int, proxies: list[str]) -> dict:
    """Insert a batch of proxies into the pool. Skips blanks and dupes within the same pool."""
    cleaned: list[str] = []
    seen: set[str] = set()
    for raw in proxies or []:
        p = str(raw or "").strip()
        if not p or p in seen:
            continue
        seen.add(p)
        cleaned.append(p)
    if not cleaned:
        return {"added": 0, "skipped": 0}
    added = 0
    skipped = 0
    with _lock:
        conn = _connect()
        for p in cleaned:
            try:
                conn.execute(
                    "INSERT INTO proxy_pool_proxies (pool_id, proxy) VALUES (?, ?)",
                    (int(pool_id), p),
                )
                added += 1
            except sqlite3.IntegrityError:
                skipped += 1
        conn.commit()
        conn.close()
    return {"added": added, "skipped": skipped}


def delete_proxy_pool_proxy(proxy_id: int) -> None:
    with _lock:
        conn = _connect()
        conn.execute("DELETE FROM proxy_pool_proxies WHERE id = ?", (int(proxy_id),))
        conn.commit()
        conn.close()


def clear_proxy_pool(pool_id: int) -> None:
    with _lock:
        conn = _connect()
        conn.execute("DELETE FROM proxy_pool_proxies WHERE pool_id = ?", (int(pool_id),))
        conn.commit()
        conn.close()


def get_proxy_pool_proxy_strings(pool_id: int) -> list[str]:
    conn = _connect()
    rows = conn.execute(
        "SELECT proxy FROM proxy_pool_proxies WHERE pool_id = ? ORDER BY id",
        (int(pool_id),),
    ).fetchall()
    conn.close()
    return [r["proxy"] for r in rows]


# ── Accounts ──────────────────────────────────────────────────────────────

def get_accounts(group_id: int | None = None) -> list[dict]:
    conn = _connect()
    if group_id is not None:
        rows = conn.execute("""
            SELECT a.*, g.name as group_name,
                   s.bearer_token, s.saved_at as session_saved_at
            FROM accounts a
            LEFT JOIN groups g ON a.group_id = g.id
            LEFT JOIN sessions s ON s.account_id = a.id
            WHERE a.group_id = ?
            ORDER BY a.name
        """, (group_id,)).fetchall()
    else:
        rows = conn.execute("""
            SELECT a.*, g.name as group_name,
                   s.bearer_token, s.saved_at as session_saved_at
            FROM accounts a
            LEFT JOIN groups g ON a.group_id = g.id
            LEFT JOIN sessions s ON s.account_id = a.id
            ORDER BY a.name
        """).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_account(account_id: int) -> dict | None:
    conn = _connect()
    row = conn.execute("""
        SELECT a.*, g.name as group_name,
               s.bearer_token, s.saved_at as session_saved_at
        FROM accounts a
        LEFT JOIN groups g ON a.group_id = g.id
        LEFT JOIN sessions s ON s.account_id = a.id
        WHERE a.id = ?
    """, (account_id,)).fetchone()
    conn.close()
    return dict(row) if row else None


def get_account_id_by_phone(phone: str) -> int | None:
    if not phone:
        return None
    conn = _connect()
    row = conn.execute("SELECT id FROM accounts WHERE phone = ?", (phone,)).fetchone()
    conn.close()
    return int(row["id"]) if row else None


def get_all_assigned_cards() -> dict[int, list[dict]]:
    """Return {account_id: [card_dict_with_label, ...]} for every assignment."""
    conn = _connect()
    rows = conn.execute("""
        SELECT a.account_id, c.*, p.name AS label
        FROM account_card_assignments a
        INNER JOIN payment_cards c ON c.id = a.card_id
        LEFT JOIN payment_pools p ON p.id = c.pool_id
        ORDER BY a.account_id, p.name, c.id
    """).fetchall()
    conn.close()
    result: dict[int, list[dict]] = {}
    for r in rows:
        d = dict(r)
        aid = int(d.pop("account_id"))
        result.setdefault(aid, []).append(d)
    return result


def add_account(
    name: str, phone: str, email: str = "",
    card_number: str = "", card_exp_month: str = "", card_exp_year: str = "", card_cvv: str = "",
    billing_name: str = "", billing_email: str = "", billing_phone: str = "",
    billing_postal: str = "", billing_country: str = "US",
    proxy: str = "", aycd_key: str = "", imap_email: str = "", imap_password: str = "", group_id: int | None = None,
) -> int:
    with _lock:
        conn = _connect()
        cur = conn.execute("""
            INSERT OR REPLACE INTO accounts
            (name, phone, email, card_number, card_exp_month, card_exp_year, card_cvv,
             billing_name, billing_email, billing_phone, billing_postal, billing_country,
             proxy, aycd_key, imap_email, imap_password, group_id)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (name, phone, email, card_number, card_exp_month, card_exp_year, card_cvv,
              billing_name, billing_email, billing_phone, billing_postal, billing_country,
              proxy, aycd_key, imap_email, imap_password, group_id))
        conn.commit()
        aid = cur.lastrowid
        conn.close()
        return aid


def update_account(account_id: int, **fields) -> None:
    """Update account fields by id."""
    if not fields: return
    allowed = {"name","phone","email","card_number","card_exp_month","card_exp_year","card_cvv",
                "billing_name","billing_email","billing_phone","billing_postal","billing_country",
                "proxy","aycd_key","imap_email","imap_password","group_id"}
    sets = [f"{k} = ?" for k in fields if k in allowed]
    vals = [fields[k] for k in fields if k in allowed]
    if not sets: return
    vals.append(account_id)
    with _lock:
        conn = _connect()
        conn.execute(f"UPDATE accounts SET {', '.join(sets)} WHERE id = ?", vals)
        conn.commit()
        conn.close()


def delete_account(account_id: int) -> None:
    with _lock:
        conn = _connect()
        conn.execute("DELETE FROM accounts WHERE id = ?", (account_id,))
        conn.commit()
        conn.close()


def assign_group(account_ids: list[int], group_id: int | None) -> None:
    with _lock:
        conn = _connect()
        conn.executemany(
            "UPDATE accounts SET group_id = ? WHERE id = ?",
            [(group_id, aid) for aid in account_ids]
        )
        conn.commit()
        conn.close()


_COUNTRY_MAP = {
    "united states": "US", "usa": "US", "canada": "CA",
    "united kingdom": "GB", "uk": "GB", "germany": "DE",
    "france": "FR", "netherlands": "NL", "australia": "AU",
    "spain": "ES", "italy": "IT", "india": "IN", "brazil": "BR",
    "mexico": "MX", "ireland": "IE", "switzerland": "CH",
}


def _col(row, *keys):
    for k in keys:
        if k in row and row[k]:
            return str(row[k]).strip()
    return ""


def _country(val):
    v = val.strip().lower()
    return _COUNTRY_MAP.get(v, val.strip().upper() if len(val.strip()) == 2 else val.strip()) or "US"


def _import_rows(rows: list[dict], group_id: int | None = None) -> int:
    """Import a list of dicts (from CSV or XLSX). Returns count of accounts touched.

    Format: one row per (account, card) pair.
    - Rows without a `card_label` column carry the account's default card (lands
      in accounts.card_number) — this matches every legacy CSV format.
    - Rows with a non-empty `card_label` create a labeled pool card and assign
      it to the account; pool is auto-created on first sighting.
    - Multiple rows for the same phone are folded into a single account: the
      first row supplies account-level fields, every row's card data is added.

    Sessions are preserved on re-import: existing accounts are UPDATEd by phone
    rather than INSERT-OR-REPLACEd, so account_id (and its FK-linked sessions
    + assigned cards) survives. Labeled-card adds are idempotent on
    (label, card_number).
    """
    from collections import OrderedDict

    by_phone: "OrderedDict[str, list[dict]]" = OrderedDict()
    for row in rows:
        cleaned = {k.strip().lower().replace(" ", "").replace("_", ""): str(v).strip() for k, v in row.items() if v}
        phone = _col(cleaned, "phone", "phonenumber")
        if not phone:
            continue
        by_phone.setdefault(phone, []).append(cleaned)

    def _billing_postal(r):
        return _col(
            r,
            "billingpostal",
            "billingpostalcode",
            "billingzip",
            "postal",
            "zip",
            "zipcode",
            "postalcode",
        )

    count = 0
    for phone, group in by_phone.items():
        # The "default-card row" is the first row (for this phone) that has no
        # card_label set. Account-level fields and the default card_number are
        # taken from it. If every row has a label, account fields come from
        # the first row and the account has no default card_number on file.
        default_row = next(
            (r for r in group if not _col(r, "cardlabel", "label")),
            group[0],
        )
        carries_default_card = not _col(default_row, "cardlabel", "label")
        email = _col(default_row, "email", "diceemail", "billingemail")

        existing_id = get_account_id_by_phone(phone)

        if existing_id is None:
            account_id = add_account(
                name=_col(default_row, "profilename", "profile", "account") or (email.split("@")[0] if email else "") or phone,
                phone=phone,
                email=email,
                card_number=_col(default_row, "cardnumber", "card") if carries_default_card else "",
                card_exp_month=_col(default_row, "cardexpmonth", "expmonth") if carries_default_card else "",
                card_exp_year=_col(default_row, "cardexpyear", "expyear") if carries_default_card else "",
                card_cvv=_col(default_row, "cvv", "cvc") if carries_default_card else "",
                billing_name=_col(default_row, "billingname", "name", "fullname"),
                billing_email=_col(default_row, "billingemail") or email,
                billing_phone=_col(default_row, "billingphone") or phone,
                billing_postal=_billing_postal(default_row),
                billing_country=_country(_col(default_row, "country", "billingcountry") or "US"),
                proxy=_col(default_row, "proxy"),
                aycd_key=_col(default_row, "aycdkey", "aycd", "aycdapikey"),
                imap_email=_col(default_row, "gmailemail", "imapemail", "otpemail") or email,
                imap_password=_col(default_row, "gmailapppassword", "imappassword", "gmailpassword", "otppassword"),
                group_id=group_id,
            )
        else:
            account_id = existing_id
            # UPSERT: only overwrite fields where the CSV has a non-empty value
            # so blank cells never wipe data the user already has on file.
            updates: dict = {}
            def _maybe(field, val):
                if val:
                    updates[field] = val

            _maybe("name", _col(default_row, "profilename", "profile", "account"))
            _maybe("email", email)
            if carries_default_card:
                _maybe("card_number", _col(default_row, "cardnumber", "card"))
                _maybe("card_exp_month", _col(default_row, "cardexpmonth", "expmonth"))
                _maybe("card_exp_year", _col(default_row, "cardexpyear", "expyear"))
                _maybe("card_cvv", _col(default_row, "cvv", "cvc"))
            _maybe("billing_name", _col(default_row, "billingname", "name", "fullname"))
            _maybe("billing_email", _col(default_row, "billingemail"))
            _maybe("billing_phone", _col(default_row, "billingphone"))
            _maybe("billing_postal", _billing_postal(default_row))
            country_raw = _col(default_row, "country", "billingcountry")
            if country_raw:
                _maybe("billing_country", _country(country_raw))
            _maybe("proxy", _col(default_row, "proxy"))
            _maybe("aycd_key", _col(default_row, "aycdkey", "aycd", "aycdapikey"))
            _maybe("imap_email", _col(default_row, "gmailemail", "imapemail", "otpemail"))
            _maybe("imap_password", _col(default_row, "gmailapppassword", "imappassword", "gmailpassword", "otppassword"))
            if group_id is not None:
                updates["group_id"] = group_id
            if updates:
                update_account(account_id, **updates)
        count += 1

        # Labeled cards: idempotent on (label, card_number) so re-importing the
        # same export is a no-op rather than a duplicate-card spree.
        existing_pairs = {
            ((c.get("label") or ""), (c.get("card_number") or ""))
            for c in (get_assigned_cards_for_account(account_id) or [])
        }
        for r in group:
            label = _col(r, "cardlabel", "label")
            card_number = _col(r, "cardnumber", "card")
            if not (label and card_number):
                continue
            if (label, card_number) in existing_pairs:
                continue
            try:
                card_id = add_payment_card(
                    label=label,
                    card_number=card_number,
                    card_exp_month=_col(r, "cardexpmonth", "expmonth"),
                    card_exp_year=_col(r, "cardexpyear", "expyear"),
                    card_cvv=_col(r, "cvv", "cvc"),
                    billing_name=_col(r, "billingname", "name", "fullname"),
                    billing_email=_col(r, "billingemail") or email,
                    billing_phone=_col(r, "billingphone") or phone,
                    billing_postal=_billing_postal(r),
                    billing_country=_country(_col(r, "country", "billingcountry") or "US"),
                )
                assign_card(account_id, card_id)
                existing_pairs.add((label, card_number))
            except Exception:
                # Skip a single bad card row but keep importing the rest.
                continue
    return count


def import_file(file_path: str, group_id: int | None = None) -> int:
    """Import accounts from CSV or XLSX. Returns count imported."""
    ext = os.path.splitext(file_path)[1].lower()

    if ext in (".xlsx", ".xls"):
        return _import_xlsx(file_path, group_id)
    else:
        return _import_csv(file_path, group_id)


def _import_csv(csv_path: str, group_id: int | None = None) -> int:
    import csv as csv_mod
    rows = []
    with open(csv_path, newline="", encoding="utf-8-sig") as f:
        reader = csv_mod.DictReader(f)
        for row in reader:
            rows.append(row)
    return _import_rows(rows, group_id)


def _import_xlsx(xlsx_path: str, group_id: int | None = None) -> int:
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h or "").strip() for h in next(rows_iter)]
    rows = []
    for values in rows_iter:
        row = {}
        for i, val in enumerate(values):
            if i < len(headers) and val is not None:
                row[headers[i]] = str(val).strip()
        if any(row.values()):
            rows.append(row)
    wb.close()
    return _import_rows(rows, group_id)


# Keep old name as alias
import_csv = import_file


# ── Sessions ──────────────────────────────────────────────────────────────

def get_session(account_id: int) -> dict | None:
    conn = _connect()
    row = conn.execute("SELECT * FROM sessions WHERE account_id = ?", (account_id,)).fetchone()
    conn.close()
    return dict(row) if row else None


def save_session(account_id: int, bearer_token: str, phone: str = "", device_id: str = "") -> None:
    with _lock:
        conn = _connect()
        conn.execute("""
            INSERT OR REPLACE INTO sessions (account_id, bearer_token, device_id, saved_at, phone)
            VALUES (?, ?, ?, ?, ?)
        """, (account_id, bearer_token, device_id, time.time(), phone))
        conn.commit()
        conn.close()


def delete_session(account_id: int) -> None:
    with _lock:
        conn = _connect()
        conn.execute("DELETE FROM sessions WHERE account_id = ?", (account_id,))
        conn.commit()
        conn.close()


def session_status(saved_at: float | None) -> str:
    """Returns 'active', 'expiring', 'expired', or 'none'."""
    if saved_at is None:
        return "none"
    age = time.time() - saved_at
    if age > SESSION_MAX_AGE:
        return "expired"
    if age > SESSION_WARN_AGE:
        return "expiring"
    return "active"


def session_expires_in(saved_at: float | None) -> str:
    """Human-readable time until session expires."""
    if saved_at is None:
        return "—"
    remaining = SESSION_MAX_AGE - (time.time() - saved_at)
    if remaining <= 0:
        return "expired"
    days = int(remaining // 86400)
    hours = int((remaining % 86400) // 3600)
    if days > 0:
        return f"{days}d {hours}h"
    return f"{hours}h"


def get_accounts_needing_auth() -> list[dict]:
    """Get accounts with no session or expired/expiring session."""
    conn = _connect()
    rows = conn.execute("""
        SELECT a.*, g.name as group_name,
               s.bearer_token, s.saved_at as session_saved_at
        FROM accounts a
        LEFT JOIN groups g ON a.group_id = g.id
        LEFT JOIN sessions s ON s.account_id = a.id
        WHERE s.account_id IS NULL
           OR (? - s.saved_at) > ?
        ORDER BY s.saved_at ASC NULLS FIRST
    """, (time.time(), SESSION_WARN_AGE)).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_accounts_with_valid_session(group_id: int | None = None) -> list[dict]:
    """Get accounts with active sessions, optionally filtered by group."""
    conn = _connect()
    if group_id is not None:
        rows = conn.execute("""
            SELECT a.*, g.name as group_name,
                   s.bearer_token, s.saved_at as session_saved_at
            FROM accounts a
            LEFT JOIN groups g ON a.group_id = g.id
            INNER JOIN sessions s ON s.account_id = a.id
            WHERE a.group_id = ?
              AND (? - s.saved_at) < ?
            ORDER BY a.name
        """, (group_id, time.time(), SESSION_MAX_AGE)).fetchall()
    else:
        rows = conn.execute("""
            SELECT a.*, g.name as group_name,
                   s.bearer_token, s.saved_at as session_saved_at
            FROM accounts a
            LEFT JOIN groups g ON a.group_id = g.id
            INNER JOIN sessions s ON s.account_id = a.id
            WHERE (? - s.saved_at) < ?
            ORDER BY a.name
        """, (time.time(), SESSION_MAX_AGE)).fetchall()
    conn.close()
    return [dict(r) for r in rows]

def account_has_valid_session(account_id: int) -> bool:
    conn = _connect()
    row = conn.execute("SELECT saved_at FROM sessions WHERE account_id = ?", (account_id,)).fetchone()
    conn.close()
    if not row:
        return False
    return (time.time() - row["saved_at"]) < SESSION_MAX_AGE


def clear_proxies_without_valid_session() -> int:
    """Clear stored account proxies unless the account has a non-expired session."""
    now = time.time()
    with _lock:
        conn = _connect()
        cur = conn.execute("""
            UPDATE accounts
            SET proxy = ''
            WHERE COALESCE(proxy, '') <> ''
              AND NOT EXISTS (
                SELECT 1
                FROM sessions s
                WHERE s.account_id = accounts.id
                  AND (? - s.saved_at) < ?
              )
        """, (now, SESSION_MAX_AGE))
        changed = cur.rowcount if cur.rowcount is not None else 0
        conn.commit()
        conn.close()
        return changed


def get_stats() -> dict:
    """Get overall account/session stats."""
    conn = _connect()
    total = conn.execute("SELECT COUNT(*) FROM accounts").fetchone()[0]
    with_session = conn.execute("""
        SELECT COUNT(*) FROM sessions WHERE (? - saved_at) < ?
    """, (time.time(), SESSION_MAX_AGE)).fetchone()[0]
    expiring = conn.execute("""
        SELECT COUNT(*) FROM sessions
        WHERE (? - saved_at) BETWEEN ? AND ?
    """, (time.time(), SESSION_WARN_AGE, SESSION_MAX_AGE)).fetchone()[0]
    groups = conn.execute("SELECT COUNT(*) FROM groups").fetchone()[0]
    conn.close()
    return {
        "total_accounts": total,
        "active_sessions": with_session,
        "expiring_sessions": expiring,
        "no_session": total - with_session,
        "groups": groups,
    }


# ── Inventory ─────────────────────────────────────────────────────────────

def get_inventory_items() -> list[dict]:
    conn = _connect()
    rows = conn.execute("""
        SELECT *
        FROM inventory_items
        ORDER BY purchased_at DESC, id DESC
    """).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def delete_inventory_item(item_id: int) -> None:
    with _lock:
        conn = _connect()
        conn.execute("DELETE FROM inventory_items WHERE id = ?", (int(item_id),))
        conn.commit()
        conn.close()


def record_inventory_purchase(
    record_key: str,
    purchase_id: str = "",
    account_id: int | None = None,
    account_name: str = "",
    account_phone: str = "",
    event_url: str = "",
    event_name: str = "",
    event_date: str = "",
    event_venue: str = "",
    ticket_type_id: str = "",
    ticket_name: str = "",
    ticket_currency: str = "USD",
    ticket_price: float | None = None,
    quantity: int = 1,
    total_price: float | None = None,
    purchase_status: str = "purchased",
    purchased_at: str | None = None,
) -> None:
    qty = max(1, int(quantity or 1))
    unit_price = float(ticket_price or 0)
    total = float(total_price) if total_price is not None else (unit_price * qty)
    stamp = purchased_at or datetime.now(UTC).isoformat().replace("+00:00", "Z")

    def _s(v) -> str:
        return "" if v is None else str(v)

    with _lock:
        conn = _connect()
        conn.execute("""
            INSERT INTO inventory_items (
                record_key, purchase_id, account_id, account_name, account_phone,
                event_url, event_name, event_date, event_venue,
                ticket_type_id, ticket_name, ticket_currency, ticket_price,
                quantity, total_price, purchase_status, purchased_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(record_key) DO UPDATE SET
                purchase_id = excluded.purchase_id,
                account_id = excluded.account_id,
                account_name = excluded.account_name,
                account_phone = excluded.account_phone,
                event_url = excluded.event_url,
                event_name = excluded.event_name,
                event_date = excluded.event_date,
                event_venue = excluded.event_venue,
                ticket_type_id = excluded.ticket_type_id,
                ticket_name = excluded.ticket_name,
                ticket_currency = excluded.ticket_currency,
                ticket_price = excluded.ticket_price,
                quantity = excluded.quantity,
                total_price = excluded.total_price,
                purchase_status = excluded.purchase_status,
                purchased_at = excluded.purchased_at
        """, (
            _s(record_key).strip(),
            _s(purchase_id).strip(),
            int(account_id) if account_id else None,
            _s(account_name).strip(),
            _s(account_phone).strip(),
            _s(event_url).strip(),
            _s(event_name).strip(),
            _s(event_date).strip(),
            _s(event_venue).strip(),
            _s(ticket_type_id).strip(),
            _s(ticket_name).strip(),
            (_s(ticket_currency) or "USD").strip().upper(),
            unit_price,
            qty,
            total,
            _s(purchase_status).strip() or "purchased",
            _s(stamp),
        ))
        conn.commit()
        conn.close()



